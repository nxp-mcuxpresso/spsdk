#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Copyright 2020-2023 NXP
#
# SPDX-License-Identifier: BSD-3-Clause
"""Module to covert Shadow register description EXCEL file to XML."""

import os
import re
import sys
from typing import Any, Dict, Optional, Tuple

import click
import openpyxl
from openpyxl import utils
from openpyxl.utils import cell as cell_utils

from spsdk.exceptions import SPSDKError
from spsdk.utils.misc import value_to_int
from spsdk.utils.registers import Registers, RegsBitField, RegsEnum, RegsRegister

XLS_COLUMN_NAMES = (
    "Block Name",
    "OTP Word",
    "Register Name",
    "Field Name",
    "Enum Name",
    "Description",
    "Shadow Register Offset/bit offset",
    "Register Width / Field width",
    "Value",
    "Access rw = has shadow register",
)


@click.group()
@click.option("-x", "--xls", type=str)
@click.option("-s", "--sheet", type=str)
@click.option("-m", "--xml", type=str)
@click.option("-t", "--xls_type", type=int, default=1)
@click.help_option("--help")
@click.pass_context
def main(ctx: click.Context, xls: str, sheet: str, xml: str, xls_type: int = 1) -> int:
    """Main CLI function."""
    if not isinstance(xls, str):
        return -1

    if not isinstance(xml, str):
        xml = ""

    ctx.obj = {"xls": xls, "sheet": sheet, "xml": xml, "xls_type": xls_type}
    return 0


@main.command()
@click.pass_obj
def convert(pass_obj: dict) -> None:
    """List supported Devices."""
    click.echo("convert")
    try:
        xls2xml_class = XLS_TYPES[str(pass_obj["xls_type"])]
        xls2xml: ShadowRegsXlsToXml = xls2xml_class(
            pass_obj["xls"], pass_obj["sheet"], pass_obj["xml"], pass_obj["xls_type"]
        )
        xls2xml.convert()
        xls2xml.registers.write_xml(xls2xml.xml_file_name)
        click.echo(f"Written XML file ({xls2xml.xml_file_name})")
        click.echo(str(xls2xml.registers))
    except Exception as exc:  # pylint: disable=broad-except
        click.echo(str(exc))


class ShadowRegsXlsToXml:
    "Class to convert XLSX to XML with shadow register description"

    def __init__(
        self, xls_file: str, sheet_name: Optional[str] = None, xml_file: str = "", xls_type: int = 1
    ) -> None:
        self.registers = Registers("Unknown")
        self.xls_type = xls_type
        self.header_cells: Dict[str, str] = {}
        self.xml_file_name = xml_file if xml_file != "" else xls_file.replace(".xlsx", ".xml")
        self.workbook = None
        self.worksheet = None
        self.sheet_name = sheet_name
        self.merged_cells = None
        click.echo(os.path.dirname(os.path.realpath(__file__)))
        self.workbook = openpyxl.load_workbook(xls_file)
        click.echo(f"Loaded XLSX file ({xls_file})")

    def convert(self) -> None:
        """Convert XLS to XML.

        :raises NotImplementedError: Derived class has to implement this method.
        """
        raise NotImplementedError("")

    def _get_worksheet(self) -> Any:
        """Find the valid worksheet with the fuse map."""
        assert self.workbook
        if self.sheet_name:
            return self.workbook[self.sheet_name]
        return self.workbook.active

    def _get_header(self) -> None:
        """Returns the dictionary with cells of header.

        :raises NotImplementedError: Derived class has to implement this method
        """
        raise NotImplementedError("Derived class has to implement this method.")

    def _get_registers(self) -> None:
        """Function finds all registers in XLS sheet and store them.

        :raises NotImplementedError: Derived class has to implement this method
        """
        raise NotImplementedError("Derived class has to implement this method.")

    def __del__(self) -> None:
        """Just close all open files."""
        if self.workbook:
            self.workbook.close()


class ShadowRegsXlsToXmlType1(ShadowRegsXlsToXml):
    """Support Type 1 XLS to convert to XML, RTxxx."""

    def convert(self) -> None:
        """Convert XLS to XML."""
        self.worksheet = self._get_worksheet()
        # Get all merged cells
        assert self.worksheet
        self.merged_cells = self.worksheet.merged_cells.ranges
        self._get_header()
        self._get_registers()

    def _get_header(self) -> None:
        """Returns the dictionary with cells of header."""
        for head in XLS_COLUMN_NAMES:
            self.header_cells[head] = self._find_cell_coor_by_val(head)

    def _filterout_bitrange(self, name: str) -> Tuple[str, bool]:  # pylint: disable=no-self-use
        """Function filter out the bit ranges in various shapes from register name."""
        bits_rev1 = re.search(r"_\d{1,4}_\d{1,4}$", name)
        bits_rev2 = re.search(r"\[\d{1,4}:\d{1,4}\]$", name)
        reverse = False
        if bits_rev1 or bits_rev2:
            bits = bits_rev1 if bits_rev1 else bits_rev2
            assert bits
            name = name.replace(bits.group(0), "")
            # Determine the order of the multiple registers.
            # Just find if the first multiple register contains zero
            bit_numbers = re.findall(r"(\d{1,4})+", bits.group(0))
            reverse = bit_numbers[len(bit_numbers) - 1] != "0"

        return name, reverse

    def _get_registers(self) -> None:
        """Function finds all registers in XLS sheet and store them."""
        assert self.worksheet
        regname_cr = cell_utils.coordinate_from_string(self.header_cells["Register Name"])
        otp_cr = cell_utils.coordinate_from_string(self.header_cells["OTP Word"])
        sr_access_cr = cell_utils.coordinate_from_string(
            self.header_cells["Access rw = has shadow register"]
        )
        desc_cr = cell_utils.coordinate_from_string(self.header_cells["Description"])
        offset_cr = cell_utils.coordinate_from_string(
            self.header_cells["Shadow Register Offset/bit offset"]
        )
        width_cr = cell_utils.coordinate_from_string(
            self.header_cells["Register Width / Field width"]
        )

        start = 1 + regname_cr[1]
        regs_group_max = 0
        regs_group_cnt = 0
        for row in range(start, self.worksheet.max_row + 1):
            cell = regname_cr[0] + str(row)
            if isinstance(self.worksheet[cell].value, str):
                # We have a register, just Mask out the Fuse register only
                access = (
                    self.worksheet[sr_access_cr[0] + str(row)].value
                    if isinstance(self.worksheet[sr_access_cr[0] + str(row)].value, str)
                    else ""
                )
                if any(x in access for x in ["rw", "ro", "wo"]):
                    # Now we have just Shadow registers only
                    # Some registers are defined multiply to store bigger data
                    # those could be detected by merged description field
                    reg_name = self.worksheet[cell].value
                    # Now, normalize the name
                    reg_name, reg_reverse = self._filterout_bitrange(reg_name)

                    reg_offset = value_to_int(self.worksheet[offset_cr[0] + str(row)].value, 16)
                    reg_width = value_to_int(self.worksheet[width_cr[0] + str(row)].value)
                    reg_descr = self.worksheet[desc_cr[0] + str(row)].value or "N/A"
                    reg_fuse_index = value_to_int(self.worksheet[otp_cr[0] + str(row)].value)
                    reg_name = reg_name.strip()

                    cells = self._get_merged_by_first_cell(desc_cr[0] + str(row))
                    if cells is not None:
                        # set the skip for next search
                        cells = cells.split(":")
                        regs_group_max = (
                            cell_utils.coordinate_from_string(cells[1])[1]
                            - cell_utils.coordinate_from_string(cells[0])[1]
                        )
                        regs_group_cnt = 0
                        reg_group_descr = reg_descr
                        reg_group_reverse = reg_reverse

                    if regs_group_max > 0:
                        reg_name = f"{reg_name}{regs_group_cnt}"
                        reg_descr = reg_group_descr
                        reg_reverse = reg_group_reverse
                        regs_group_cnt += 1
                        if regs_group_cnt > regs_group_max:
                            regs_group_max = 0
                            regs_group_cnt = 0
                    print(f"Reg: {reg_name}")
                    reg_descr = reg_descr.replace("\n", "&#10;")
                    register = RegsRegister(
                        reg_name,
                        reg_offset,
                        reg_width,
                        reg_descr,
                        reg_reverse,
                        access,
                        otp_index=reg_fuse_index,
                    )

                    self.registers.add_register(register)

                    cells = self._get_merged_by_first_cell(regname_cr[0] + str(row))
                    if cells is not None:
                        # find the number of rows of the register description
                        cells = cells.split(":")
                        reg_lines = (
                            cell_utils.coordinate_from_string(cells[1])[1]
                            - cell_utils.coordinate_from_string(cells[0])[1]
                        )
                    self._get_bitfields(register, row, reg_lines + 1)

    def _get_bitfields(self, reg: Any, excel_row: int, excel_row_cnt: int) -> None:
        """Tried to find and fill up all register bitfields."""
        assert self.worksheet
        if excel_row_cnt <= 1:
            # There is no bitfields
            return

        bitfieldname_cr = cell_utils.coordinate_from_string(self.header_cells["Field Name"])
        desc_cr = cell_utils.coordinate_from_string(self.header_cells["Description"])
        offset_cr = cell_utils.coordinate_from_string(
            self.header_cells["Shadow Register Offset/bit offset"]
        )
        width_cr = cell_utils.coordinate_from_string(
            self.header_cells["Register Width / Field width"]
        )
        rv_cr = cell_utils.coordinate_from_string(self.header_cells["Value"])

        excel_row += 1
        excel_row_cnt -= 1

        for row in range(excel_row, excel_row + excel_row_cnt):
            cell = bitfieldname_cr[0] + str(row)
            if isinstance(self.worksheet[cell].value, str):
                try:
                    bitfield_name = self.worksheet[cell].value
                    bitfield_offset = value_to_int(self.worksheet[offset_cr[0] + str(row)].value)
                    bitfield_width = value_to_int(self.worksheet[width_cr[0] + str(row)].value)
                    bitfield_descr = self.worksheet[desc_cr[0] + str(row)].value or "N/A"
                    bitfield_rv = self.worksheet[rv_cr[0] + str(row)].value or "N/A"
                    bitfield_descr = bitfield_descr.replace("\n", "&#10;")
                    print(f"  Bitfield: {bitfield_name}")
                    bitfield = RegsBitField(
                        reg,
                        bitfield_name,
                        bitfield_offset,
                        bitfield_width,
                        bitfield_descr,
                        reset_val=bitfield_rv,
                    )
                    reg.add_bitfield(bitfield)
                except Exception as exc:
                    print(f"Error raised during loading bitfield {bitfield_name}. {exc}")
                    raise SPSDKError(str(exc)) from exc
                cells = self._get_merged_by_first_cell(bitfieldname_cr[0] + str(row))
                if cells is not None:
                    # find the number of rows of the register description
                    cells = cells.split(":")
                    reg_lines = (
                        cell_utils.coordinate_from_string(cells[1])[1]
                        - cell_utils.coordinate_from_string(cells[0])[1]
                    )
                    self._get_enums(bitfield, row, reg_lines + 1)

    def _get_enums(self, bitfield: Any, excel_row: int, excel_row_cnt: int) -> None:
        """Tried to find and fill up all register bitfields enumerations."""
        assert self.worksheet
        if excel_row_cnt <= 1:
            # There is no enums
            return

        enum_name_cr = cell_utils.coordinate_from_string(self.header_cells["Enum Name"])
        desc_cr = cell_utils.coordinate_from_string(self.header_cells["Description"])
        value_cr = cell_utils.coordinate_from_string(self.header_cells["Value"])

        excel_row += 1
        excel_row_cnt -= 1

        for row in range(excel_row, excel_row + excel_row_cnt):
            cell = enum_name_cr[0] + str(row)

            if isinstance(self.worksheet[cell].value, str):
                try:
                    enum_name = self.worksheet[cell].value
                    enum_descr = self.worksheet[desc_cr[0] + str(row)].value or "N/A"
                    enum_value: str = self.worksheet[value_cr[0] + str(row)].value
                    enum_value = enum_value.replace("b'", "0b")
                    enum_descr = enum_descr.replace("\n", "&#10;")
                    print(f"    Enum: {enum_name}")
                    if enum_value is None:
                        click.echo(
                            f"Warning: The Enum {enum_name} is missing and it will be skipped."
                        )
                    else:
                        bitfield.add_enum(
                            RegsEnum(enum_name, enum_value, enum_descr, bitfield.width)
                        )
                except Exception as exc:
                    print(f"Error raised during loading enum {enum_name}. {exc}")
                    raise SPSDKError(str(exc)) from exc

    def _get_merged_by_first_cell(self, cell: str) -> str:
        """Function returns the merged range by first cell."""
        assert self.merged_cells
        for merged in self.merged_cells:
            if merged.coord.find(cell + ":") >= 0:
                return merged.coord
        return None

    def _find_cell_coor_by_val(self, value: Any, start: str = "", end: str = "") -> str:
        """Search engine for the cell values"""
        assert self.worksheet
        if start is None or start == "":
            start = "A1"
        if end is None or end == "":
            end = utils.get_column_letter(self.worksheet.max_column) + str(self.worksheet.max_row)

        start_cell = cell_utils.coordinate_from_string(start)
        end_cell = cell_utils.coordinate_from_string(end)
        start_column = utils.column_index_from_string(start_cell[0])
        start_row = start_cell[1]
        end_column = utils.column_index_from_string(end_cell[0])
        end_row = end_cell[1]

        for row in range(start_row, end_row + 1):
            for column in range(start_column, end_column + 1):
                val = self.worksheet[utils.get_column_letter(column) + str(row)].value
                if isinstance(val, str):
                    val = val.replace("\n", " ")
                    val = val.replace("  ", " ")
                if value == val:
                    return utils.get_column_letter(column) + str(row)

        return None


class ShadowRegsXlsToXmlType2(ShadowRegsXlsToXml):
    """Convert Type2 of XLS (RTxxxx)."""

    def convert(self) -> None:
        assert self.worksheet
        self.sheet_name = "Fuse Definitions"
        self.worksheet = self._get_worksheet()
        # Get all merged cells
        self._get_header()
        self._get_registers()

    def _get_header(self) -> None:
        """Returns header of sheet."""
        self.header_cells["reg_base"] = "A"
        self.header_cells["fuse_address"] = "B"
        self.header_cells["fuse_index"] = "D"
        self.header_cells["fuse_name"] = "E"
        self.header_cells["fuse_width"] = "F"
        self.header_cells["fuse_descr"] = "G"
        self.header_cells["fuse_sett"] = "H"
        self.header_cells["burned_value"] = "J"
        self.header_cells["customer_visible"] = "L"

    def _get_regbase(self, line: int) -> int:
        """Return  register base address."""
        assert self.worksheet
        return int(self.worksheet[self.header_cells["reg_base"] + str(line)].value, 16)

    def _get_fusename(self, line: int) -> int:
        """Return Fuse name."""
        assert self.worksheet
        try:
            name = self.worksheet[self.header_cells["fuse_name"] + str(line)].value
        except Exception:  # pylint: disable=broad-except
            name = "Unknown name :-("
        return name

    def _get_fuseoffset(self, line: int) -> int:
        """Gets Fuse offset."""
        assert self.worksheet
        reg_offset_bits = self._get_regbase(line) - 0x400
        fuse_offset = self.worksheet[self.header_cells["fuse_index"] + str(line)].value
        fuse_offset = fuse_offset - reg_offset_bits
        return fuse_offset

    def _get_fusewidth(self, line: int) -> int:
        """Get fuse width."""
        assert self.worksheet
        return self.worksheet[self.header_cells["fuse_width"] + str(line)].value

    def _get_fusedescription(self, line: int) -> int:
        """Return fuse description."""
        assert self.worksheet
        try:
            fuse_description = self.worksheet[self.header_cells["fuse_descr"] + str(line)].value
        except Exception:  # pylint: disable=broad-except
            fuse_description = "There is no any special description"
        return fuse_description

    def _get_fuse_resetvalue(self, line: int) -> int:
        """Return fuse reset value."""
        assert self.worksheet
        try:
            fuse_resetvalue = self.worksheet[self.header_cells["burned_value"] + str(line)].value
        except Exception:  # pylint: disable=broad-except
            fuse_resetvalue = "0"
        return fuse_resetvalue

    def _get_fuse_bitfield_info(self, line: int) -> Tuple[int, int]:
        """Return Fuse bitfield information."""
        assert self.worksheet
        try:
            fuse_width = self._get_fusewidth(line)
            fuse_address = self.worksheet[self.header_cells["fuse_address"] + str(line)].value
            pattern = re.compile(r"\[([^)]*)\]")
            offsets = pattern.findall(fuse_address)[0]
            if offsets.count(":") > 0:
                offsets = offsets.split(":")
                offsets.reverse()
                offset = int(offsets[0])
            else:
                offset = int(offsets)
        except Exception as exc:  # pylint: disable=broad-except
            click.echo(f"Issue with get the getting bitfield info ({str(exc)})")

        return offset, fuse_width

    def _get_registers(self) -> None:
        """Return all registers from XLS file."""
        assert self.worksheet
        # Start line in excel style 2 is 3!
        reg_base = 0
        try:
            for row in range(3, self.worksheet.max_row + 1):
                new_reg_base = self._get_regbase(row)
                if new_reg_base != reg_base:
                    # This is new register, just create it
                    reg_base = new_reg_base
                    reg_name = f"REG_0x{reg_base:04X}"
                    reg_offset = 0x400 - reg_base
                    reg_width = 32  # TODO solve that fields
                    reg_description = f"This is description string of {reg_name} register"
                    reg_reverse = False
                    reg_access = "RW"
                    reg = RegsRegister(
                        reg_name, reg_offset, reg_width, reg_description, reg_reverse, reg_access
                    )
                    self.registers.add_register(reg)

                # we have added register, so this is about a adding of bitfield
                bitfield_name = self._get_fusename(row)
                bitfield_offset, bitfield_width = self._get_fuse_bitfield_info(row)
                bitfield_descr = self._get_fusedescription(row)
                bitfield_rv = self._get_fuse_resetvalue(row)
                bitfield = RegsBitField(
                    reg,
                    bitfield_name,
                    bitfield_offset,
                    bitfield_width,
                    bitfield_descr,
                    reset_val=bitfield_rv,
                )
                reg.add_bitfield(bitfield)
        except Exception as exc:  # pylint: disable=broad-except
            click.echo(f"Unwanted exception during getting registers({str(exc)})")


XLS_TYPES = {"1": ShadowRegsXlsToXmlType1, "2": ShadowRegsXlsToXmlType2}

if __name__ == "__main__":
    sys.exit(main())  # pragma: no cover # pylint: disable=no-value-for-parameter
#     xls = ShadowRegsXlsToXml("tools/OTP6.xlsx")
# regs = Registers("pokus 685", None)
# regs.load_registers_from_xml("tools/OTP.xml")
# regs.write_xml("tools/OPT2.xml")
